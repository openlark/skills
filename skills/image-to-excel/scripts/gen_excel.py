#!/usr/bin/env python3
"""
Image to Table — Excel Generation Script
Usage: python3 gen_excel.py <output_path> '<json_data>'
json_data: 2D array JSON, e.g., '[["Name","Age"],["Zhang San",25],["Li Si",30]]'
"""
import sys
import json

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Missing openpyxl, installing...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter


def generate_excel(output_path: str, rows: list[list]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Table"

    # Style definitions
    header_font = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_font = Font(name="Microsoft YaHei", size=11)
    cell_align = Alignment(vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )

    # Write data
    has_header = len(rows) > 1
    for r_idx, row in enumerate(rows):
        for c_idx, val in enumerate(row):
            cell = ws.cell(row=r_idx + 1, column=c_idx + 1, value=str(val) if val else "")
            cell.border = thin_border
            cell.alignment = cell_align
            if r_idx == 0 and has_header:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            else:
                cell.font = cell_font

    # Auto-fit column width
    for c_idx in range(1, len(rows[0]) + 1):
        max_len = 0
        for row in rows:
            val = str(row[c_idx - 1]) if row[c_idx - 1] else ""
            # Chinese characters count as 2 width units
            char_len = sum(2 if ord(ch) > 127 else 1 for ch in val)
            max_len = max(max_len, char_len)
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 4, 40)

    # Freeze header row
    if has_header:
        ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"✅ Excel saved: {output_path} ({len(rows)} rows × {len(rows[0])} columns)")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python3 gen_excel.py <output_path> '<json_data>'")
        sys.exit(1)

    output_path = sys.argv[1]
    json_str = sys.argv[2]

    try:
        rows = json.loads(json_str)
    except json.JSONDecodeError as e:
        print(f"❌ JSON parsing failed: {e}")
        sys.exit(1)

    if not isinstance(rows, list) or not all(isinstance(r, list) for r in rows):
        print("❌ Invalid data format, 2D array required")
        sys.exit(1)

    # Pad inconsistent rows
    max_cols = max(len(r) for r in rows) if rows else 0
    for row in rows:
        while len(row) < max_cols:
            row.append("")

    generate_excel(output_path, rows)